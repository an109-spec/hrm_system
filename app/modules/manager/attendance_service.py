from calendar import calendar
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import extract, or_
from app.extensions import db
from app.models.attendance import Attendance, AttendanceShiftStatus, AttendanceType
from app.models.employee import Employee
from app.models.file_upload import FileUpload
from app.models.leave import LeaveRequest
from app.models.notification import Notification
from app.models.overtime_request import OvertimeRequest
from app.modules.attendance.service import AttendanceService
from app.utils.time import get_current_time
from datetime import date, datetime, timedelta
from app.utils.upload_service import UploadService
from app.modules.leave.service import LeaveService
class AttendanceManagerService:
    @staticmethod
    def _get_subordinates(manager_id: int) -> list[Employee]:
        """
        Lấy toàn bộ nhân viên thuộc (các) phòng ban mà manager này quản lý.
        Ràng buộc: Chỉ lấy nhân viên trong cùng phòng ban.
        """
        manager = Employee.query.get(manager_id)
        if not manager or manager.is_deleted:
            return []
        managed_depts = manager.managed_department
        if not managed_depts:
            return []
        if not isinstance(managed_depts, list):
            managed_depts = [managed_depts]
        all_subordinates = []
        for dept in managed_depts:
            if hasattr(dept, 'employees'):
                emps = [
                    e for e in dept.employees 
                    if e.id != manager_id and not e.is_deleted
                ]
                all_subordinates.extend(emps)
        return all_subordinates
    
    @staticmethod
    def get_department_attendance_rows(manager_id: int, filters: dict | None = None, paginate: bool = True) -> dict:
        filters = filters or {}
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        if not subordinates:
            return {"items": [], "total_pages": 0, "current_page": 1, "total_items": 0}
        subordinate_ids = [e.id for e in subordinates]
        subordinate_map = {e.id: e for e in subordinates}
        query = Attendance.query.filter(
            Attendance.employee_id.in_(subordinate_ids),
            extract('month', Attendance.date) == int(filters.get("month", get_current_time().month)),
            extract('year', Attendance.date) == int(filters.get("year", get_current_time().year))
        )
        status_filter = filters.get("status", "").strip().lower()
        if status_filter:
            query = query.filter(Attendance.shift_status == status_filter)
        query = query.order_by(Attendance.date.desc(), Attendance.employee_id.asc())
        if paginate:
            page = int(filters.get("page", 1))
            per_page = int(filters.get("per_page", 10))
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)
            items_to_process = pagination.items
            total_pages = pagination.pages
            total_items = pagination.total
            current_page = pagination.page
        else:
            items_to_process = query.all()
            total_pages = 1
            total_items = len(items_to_process)
            current_page = 1
        page_employee_ids = list(set(att.employee_id for att in items_to_process))
        approved_leaves = LeaveRequest.query.filter(
            LeaveRequest.employee_id.in_(page_employee_ids),
            LeaveRequest.status == 'approved',
            LeaveRequest.is_deleted.is_(False)
        ).all()
        leave_lookup = {}
        for leave in approved_leaves:
            curr_d = leave.from_date
            while curr_d <= leave.to_date:
                leave_lookup[(leave.employee_id, curr_d)] = True
                curr_d += timedelta(days=1)
        data_items = []
        for att in items_to_process:
            emp = subordinate_map.get(att.employee_id)
            if not emp: continue
            is_on_leave = leave_lookup.get((att.employee_id, att.date), False)
            final_status = AttendanceShiftStatus.LEAVE if is_on_leave else att.normalized_shift_status
            data_items.append({
                "employee_id": emp.id,
                "name": emp.full_name,
                "department": emp.department.name if emp.department else "--",
                "date": att.date.strftime("%d/%m/%Y"),
                "check_in": att.check_in.strftime("%H:%M") if att.check_in else "--:--",
                "check_out": att.check_out.strftime("%H:%M") if att.check_out else "--:--",
                "status_label": AttendanceShiftStatus.label(final_status),
                "worked_hours": float(att.working_hours or 0),
                "overtime_hours": float(att.overtime_hours or 0),
                "notes": att.notes
            })
        return {
            "items": data_items,
            "total_pages": total_pages,
            "current_page": current_page,
            "total_items": total_items
        }
        
    @staticmethod
    def export_attendance_report(manager_id: int, filters: dict | None = None) -> dict:
        data = AttendanceManagerService.get_department_attendance_rows(manager_id, filters, paginate=False)
        items = data.get("items", [])
        wb = Workbook()
        ws = wb.active
        ws.title = "BaoCaoChamCong"
        headers = ["STT", "Mã NV", "Họ và Tên", "Phòng ban", "Ngày", "Vào", "Ra", "Trạng thái", "Giờ làm", "Tăng ca", "Ghi chú"]
        ws.append(headers)
        # Định dạng header
        fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill, cell.font, cell.alignment = fill, font, Alignment(horizontal="center")
        for idx, item in enumerate(items, start=1):
            ws.append([idx, item["employee_id"], item["name"], item["department"], item["date"], 
                    item["check_in"], item["check_out"], item["status_label"], 
                    item["worked_hours"], item["overtime_hours"], item["notes"]])
        # 3. Sử dụng UploadService để chuẩn hóa file lưu
        entity_type = "reports"
        UploadService._ensure_folder(entity_type)
        # Tạo tên file duy nhất (UUID) thay vì thời gian để tránh trùng lặp
        filename = UploadService._generate_filename("report.xlsx")
        relative_path = os.path.join(entity_type, filename)
        full_path = os.path.join(UploadService.UPLOAD_FOLDER, relative_path)
        wb.save(full_path)
        # 4. Lưu vào Database
        manager_user_id = Employee.query.get(manager_id).user_id
        file_rec = FileUpload(
            file_name=f"Report_{get_current_time().strftime('%Y%m%d')}.xlsx", # Tên người dùng thấy
            file_url=relative_path, # Đường dẫn hệ thống
            file_type="document",
            file_size=os.path.getsize(full_path),
            uploaded_by=manager_user_id,
            entity_type=entity_type,
            entity_id=manager_id
        )
        db.session.add(file_rec)
        db.session.commit()
        return {"file_url": f"/static/uploads/{relative_path}"}
    
    @staticmethod
    def get_department_attendance_summary(manager_id: int, filters: dict | None = None) -> dict:
        filters = filters or {}
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        if not subordinates:
            return {
                "total_employees": 0, "checked_in": 0, "not_checked_in": 0,
                "on_leave": 0, "late": 0, "overtime": 0
            }
        subordinate_ids = [e.id for e in subordinates]
        all_rows = Attendance.query.filter(
            Attendance.employee_id.in_(subordinate_ids),
            extract('month', Attendance.date) == int(filters.get("month", get_current_time().month)),
            extract('year', Attendance.date) == int(filters.get("year", get_current_time().year))
        ).all()
        return {
            "total_employees": len(set(x.employee_id for x in all_rows)),
            "checked_in": len([x for x in all_rows if x.check_in is not None]),
            "not_checked_in": len([x for x in all_rows if x.check_in is None and x.normalized_shift_status != AttendanceShiftStatus.LEAVE]),
            "on_leave": len([x for x in all_rows if x.normalized_shift_status == AttendanceShiftStatus.LEAVE]),
            "late": len([x for x in all_rows if x.late_minutes > 0]), 
            "overtime": len([x for x in all_rows if float(x.overtime_hours or 0) > 0]),
        }
    
    @staticmethod
    def get_department_attendance_detail(manager_id: int, employee_id: int, filters: dict | None = None) -> dict:
        """
        Xem chi tiết trạng thái điểm danh, đơn từ của nhân viên trong ngày.
        """
        # 1. Kiểm tra quyền
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        subordinate_ids = [e.id for e in subordinates]
        if employee_id not in subordinate_ids:
            raise ValueError("Bạn không có quyền xem thông tin của nhân viên này.")
        
        now = get_current_time()
        today = now.date()

        # 2. Lấy thông tin nhân viên
        employee = Employee.query.get(employee_id)
        if not employee or employee.is_deleted:
            raise ValueError("Nhân viên không tồn tại hoặc đã nghỉ việc.")

        # Lấy bản ghi điểm danh
        att_record = Attendance.query.filter_by(employee_id=employee_id, date=today).first()

        # 3. Sử dụng Service để lấy đơn nghỉ phép (Đã refactor để tránh trùng lặp)
        leave = LeaveService.get_active_leave_on_date(employee_id, today)

        # 4. Lấy đơn tăng ca
        ot_request = OvertimeRequest.query.filter(
            OvertimeRequest.employee_id == employee_id,
            OvertimeRequest.overtime_date == today,
            OvertimeRequest.is_deleted.is_(False)
        ).first()

        # 5. Lấy lịch sử 5 ngày gần nhất
        history_rows = Attendance.query.filter(
            Attendance.employee_id == employee_id,
            Attendance.date < today
        ).order_by(Attendance.date.desc()).limit(5).all()

        # 6. Build Data trả về
        return {
            "employee_info": {
                "id": employee.id,
                "full_name": employee.full_name,
                "department": employee.department.name if employee.department else "--"
            },
            "today_summary": {
                "date": today.strftime("%d/%m/%Y"),
                "check_in": att_record.check_in.strftime("%H:%M:%S") if att_record and att_record.check_in else None,
                "check_out": att_record.check_out.strftime("%H:%M:%S") if att_record and att_record.check_out else None,
                "shift_status": att_record.shift_status if att_record else AttendanceShiftStatus.NOT_STARTED,
                "status_label": att_record.shift_status_label if att_record else "Chưa bắt đầu",
                "attendance_type_label": att_record.attendance_type_label if att_record else "Ngày thường",
                "worked_hours": float(att_record.working_hours or 0) if att_record else 0,
                "overtime_hours": float(att_record.overtime_hours or 0) if att_record else 0,
                "is_abnormal": bool(att_record and att_record.attendance_type == AttendanceType.ABNORMAL)
            },
            "leave_info": {
                "type": leave.leave_type.name if leave and leave.leave_type else None,
                "duration": f"{leave.from_date} tới {leave.to_date}" if leave else None,
                "reason": leave.reason if leave else None
            } if leave else None,
            "ot_info": {
                "requested_hours": float(ot_request.requested_hours or 0),
                "status": ot_request.status,
                "reason": ot_request.reason
            } if ot_request else None,
            "recent_history": [
                {
                    "date": h.date.strftime("%d/%m/%Y"),
                    "status": h.shift_status_label,
                    "hours": float(h.working_hours or 0),
                    "ot": float(h.overtime_hours or 0)
                }
                for h in history_rows
            ]
        }
    
    @staticmethod
    def auto_detect_abnormal_records(manager_id: int):
        now = get_current_time()
        today = now.date()
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        subordinate_ids = [e.id for e in subordinates]
        if not subordinate_ids:
            return
        query_forgot = Attendance.query.filter(
            Attendance.employee_id.in_(subordinate_ids),
            Attendance.check_in.isnot(None),
            Attendance.check_out.is_(None),
            Attendance.attendance_type != Attendance.Type.ABNORMAL,
            db.or_(
                Attendance.date < today,
                db.and_(Attendance.date == today, now.hour >= 18)
            )
        )
        
        for rec in query_forgot.all():
            rec.set_attendance_type(Attendance.Type.ABNORMAL)
            rec.set_shift_status(Attendance.ShiftStatus.REGULAR_CHECKOUT_REQUIRED)
            rec.notes = "Hệ thống: Quá giờ làm việc nhưng thiếu dữ liệu check-out."
        illogical_records = Attendance.query.filter(
            Attendance.employee_id.in_(subordinate_ids),
            Attendance.check_in.isnot(None),
            Attendance.check_out.isnot(None),
            Attendance.check_out < Attendance.check_in,
            Attendance.attendance_type != Attendance.Type.ABNORMAL
        ).all()

        for rec in illogical_records:
            rec.set_attendance_type(Attendance.Type.ABNORMAL)
            rec.notes = "Hệ thống: Giờ ra sớm hơn giờ vào (Dữ liệu xung đột)."

        # 3. Xử lý Vắng mặt (Chạy sau 10h sáng)
        if now.hour >= 10:
            checked_in_ids = {
                a.employee_id for a in Attendance.query.filter_by(date=today).filter(
                    Attendance.employee_id.in_(subordinate_ids)
                ).all()
            }
            on_leave_ids = AttendanceManagerService._approved_leave_employee_ids(subordinate_ids, today)
            absent_ids = set(subordinate_ids) - checked_in_ids - on_leave_ids
            for emp_id in absent_ids:
                new_absent_record = Attendance(
                    employee_id=emp_id,
                    date=today,
                    notes="Hệ thống: Vắng mặt không lý do (Không check-in & không có đơn nghỉ phép)."
                )
                new_absent_record.set_attendance_type(Attendance.Type.ABNORMAL)
                new_absent_record.set_shift_status(Attendance.ShiftStatus.ABSENT)
                db.session.add(new_absent_record)
        db.session.commit()

    @staticmethod
    def update_attendance_correction(manager_id: int, attendance_id: int, 
                                    new_check_in: datetime = None, 
                                    new_check_out: datetime = None, 
                                    reason: str = "") -> dict:
        record = Attendance.query.get(attendance_id)
        if not record:
            raise ValueError("Không tìm thấy dữ liệu chấm công.")
        subordinate_ids = {e.id for e in AttendanceManagerService._get_subordinates(manager_id)}
        if record.employee_id not in subordinate_ids:
            raise ValueError("Bạn không có quyền chỉnh sửa nhân viên này.")
        if new_check_in: record.check_in = new_check_in
        if new_check_out: record.check_out = new_check_out
        if record.check_in and record.check_out:
            if record.check_out < record.check_in:
                raise ValueError("Lỗi logic: Giờ check-out không thể sớm hơn check-in.")
            record.set_attendance_type(Attendance.Type.NORMAL)
            record.set_shift_status(Attendance.ShiftStatus.COMPLETED)
        elif record.check_in and not record.check_out:
            record.set_attendance_type(Attendance.Type.ABNORMAL)
            record.set_shift_status(Attendance.ShiftStatus.REGULAR_CHECKOUT_REQUIRED)
        now = get_current_time()
        timestamp_str = now.strftime('%Y-%m-%d %H:%M')
        record.notes = f"[Manager Edited]: {reason} - Updated at {timestamp_str}"
        db.session.commit()
        return {
            "status": "success",
            "message": "Dữ liệu đã được cập nhật.",
            "new_status_label": record.shift_status_label
        }
        
    @staticmethod
    def send_reminder(manager_id: int, employee_ids: list[int], message: str | None = None) -> bool:
        """
        Gửi thông báo nhắc nhở chấm công cho danh sách nhân viên.
        CHỈ gửi cho những nhân viên thuộc quyền quản lý của manager_id.
        """
        if not employee_ids:
            return False
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        valid_subordinate_ids = {e.id for e in subordinates}
        authorized_ids = [emp_id for emp_id in employee_ids if emp_id in valid_subordinate_ids]
        if not authorized_ids:
            return False 
        employees = Employee.query.filter(
            Employee.id.in_(authorized_ids),
            Employee.is_deleted.is_(False),
            Employee.user_id.isnot(None) 
        ).all()
        if not employees:
            return False
        for emp in employees:
            final_message = message if message else f"Chào {emp.full_name}, bạn chưa thực hiện chấm công hôm nay. Đừng quên nhé!"
            new_notification = Notification(
                user_id=emp.user_id,
                title="🔔 Nhắc nhở chấm công",
                content=final_message,
                type="reminder",
                link="/employee/attendance", 
                is_read=False
            )
            db.session.add(new_notification)
        try:
            db.session.commit()
            return True
        except Exception as e:
            db.session.rollback()
            return False
        
    @staticmethod
    def get_manager_attendance_dashboard(manager_id: int, month: int, year: int) -> list[dict]:
        """
        Dashboard thống kê chấm công đơn giản.
        Chỉ trả về số liệu thống kê (stats) của nhân viên.
        Không xử lý các bản ghi bất thường.
        """
        first_day = date(year, month, 1)
        _, last_day_num = calendar.monthrange(year, month)
        last_day = date(year, month, last_day_num)
        subordinates = AttendanceManagerService._get_subordinates(manager_id)
        if not subordinates:
            return []
        sub_ids = [e.id for e in subordinates]
        attendances = Attendance.query.filter(
            Attendance.employee_id.in_(sub_ids),
            Attendance.date >= first_day,
            Attendance.date <= last_day
        ).all()
        leaves = LeaveRequest.query.filter(
            LeaveRequest.employee_id.in_(sub_ids),
            LeaveRequest.status == 'approved', 
            LeaveRequest.is_deleted.is_(False),
            or_(
                LeaveRequest.from_date.between(first_day, last_day),
                LeaveRequest.to_date.between(first_day, last_day),
                (LeaveRequest.from_date <= first_day) & (LeaveRequest.to_date >= last_day)
            )
        ).all()
        results = []
        for emp in subordinates:
            emp_atts = [a for a in attendances if a.employee_id == emp.id]
            emp_leaves = [l for l in leaves if l.employee_id == emp.id]
            stats = {"on_time": 0, "late": 0, "absent": 0, "leave": 0}
            for att in emp_atts:
                status = AttendanceService.get_attendance_status(att)
                if status in stats:
                    stats[status] += 1
            for lv in emp_leaves:
                overlap_start = max(lv.from_date, first_day)
                overlap_end = min(lv.to_date, last_day)
                days = (overlap_end - overlap_start).days + 1
                stats["leave"] += max(0, days)
            results.append({
                "employee_id": emp.id,
                "full_name": emp.full_name,
                "stats": stats
            })
        return results